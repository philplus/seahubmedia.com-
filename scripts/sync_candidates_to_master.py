#!/usr/bin/env python3
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def norm_phone(p: str) -> str:
    if not p:
        return ""
    s = str(p).strip().replace(" ", "")
    # keep leading + if present
    return s


def main():
    if len(sys.argv) < 4:
        print("Usage: sync_candidates_to_master.py <batch_xlsx> <master_xlsx> <default_role>")
        sys.exit(2)

    batch_path = Path(sys.argv[1])
    master_path = Path(sys.argv[2])
    default_role = sys.argv[3]

    wb_batch = load_workbook(batch_path)
    ws_b = wb_batch.active

    wb_master = load_workbook(master_path)
    ws_m = wb_master.active

    # master headers (row 1)
    headers = [ws_m.cell(1, c).value for c in range(1, ws_m.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    required = ["岗位", "姓名", "手机", "WhatsApp", "LinkedIn"]
    for r in required:
        if r not in col:
            raise RuntimeError(f"Master missing column: {r}")

    # existing phones for dedup
    existing = set()
    for r in range(2, ws_m.max_row + 1):
        existing.add(norm_phone(ws_m.cell(r, col["手机"]).value))
        existing.add(norm_phone(ws_m.cell(r, col["WhatsApp"]).value))

    # batch header map
    b_headers = [ws_b.cell(1, c).value for c in range(1, ws_b.max_column + 1)]
    bcol = {h: i + 1 for i, h in enumerate(b_headers) if h}

    now = datetime.now().isoformat(timespec="seconds")
    added = 0
    skipped = 0

    for r in range(2, ws_b.max_row + 1):
        name = ws_b.cell(r, bcol.get("Name", 0)).value if bcol.get("Name") else None
        phone = ws_b.cell(r, bcol.get("Phone", 0)).value if bcol.get("Phone") else None
        url = ws_b.cell(r, bcol.get("Profile URL", 0)).value if bcol.get("Profile URL") else None
        source = ws_b.cell(r, bcol.get("Source", 0)).value if bcol.get("Source") else None

        phone_n = norm_phone(phone)
        if not name and not phone_n:
            continue
        if phone_n and phone_n in existing:
            skipped += 1
            continue

        new_row = ws_m.max_row + 1
        ws_m.cell(new_row, col["岗位"]).value = default_role
        ws_m.cell(new_row, col["姓名"]).value = name
        ws_m.cell(new_row, col["手机"]).value = phone_n
        ws_m.cell(new_row, col["WhatsApp"]).value = phone_n
        if "LinkedIn" in col:
            ws_m.cell(new_row, col["LinkedIn"]).value = url

        # optional columns if present
        if "来源" in col:
            ws_m.cell(new_row, col["来源"]).value = source
        if "备注" in col:
            ws_m.cell(new_row, col["备注"]).value = f"imported {now} from {batch_path.name}"

        added += 1
        if phone_n:
            existing.add(phone_n)

    wb_master.save(master_path)
    print(f"added={added} skipped={skipped} master_rows={ws_m.max_row}")


if __name__ == "__main__":
    main()
