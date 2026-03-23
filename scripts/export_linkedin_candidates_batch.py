#!/usr/bin/env python3
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def autosize(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def main():
    if len(sys.argv) < 3:
        print("Usage: export_linkedin_candidates_batch.py <input_json_path> <output_xlsx_path>")
        sys.exit(2)

    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2])

    data = json.loads(in_path.read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "candidates"

    headers = [
        "No",
        "Role",
        "Name",
        "Title",
        "Location",
        "LinkedIn",
        "Experience Check",
        "Source Query",
        "Exported At",
    ]
    ws.append(headers)

    exported_at = datetime.now().isoformat(timespec="seconds")

    for i, r in enumerate(data, start=1):
        ws.append([
            i,
            r.get("role"),
            r.get("name"),
            r.get("title"),
            r.get("location"),
            r.get("url"),
            r.get("experienceCheck"),
            r.get("sourceQuery"),
            exported_at,
        ])

    autosize(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(str(out_path))


if __name__ == "__main__":
    main()
