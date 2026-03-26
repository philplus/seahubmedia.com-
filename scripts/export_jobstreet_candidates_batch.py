#!/usr/bin/env python3
import re
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

STATUS = ["Baru", "Kotak masuk", "Terpilih", "Tidak cocok"]

def parse_row(row_text: str):
    t = row_text or ""
    status = next((s for s in STATUS if s in t), "")
    applied_for = ""
    m = re.search(r"Lamaran untuk\s*(.+?)\s*(\d+\s+hari\s+yang\s+lalu|\d+\s+minggu\s+yang\s+lalu|\d+\s+bulan\s+yang\s+lalu)", t)
    if m:
        applied_for = m.group(1).strip()
    applied_age = ""
    m2 = re.search(r"(\d+\s+hari\s+yang\s+lalu|\d+\s+minggu\s+yang\s+lalu|\d+\s+bulan\s+yang\s+lalu)", t)
    if m2:
        applied_age = m2.group(1)

    current_title = ""
    current_company = ""
    if status and "Lamaran untuk" in t:
        mid = t.split(status, 1)[1]
        mid = mid.split("Lamaran untuk", 1)[0]
        mid = re.sub(r"\+?\d[\d\s]{7,}", " ", mid)
        mid = re.sub(r"\s+", " ", mid).strip()
        if " di " in mid:
            left, right = mid.split(" di ", 1)
            current_title = left.strip()
            current_company = right.strip()
        else:
            parts = mid.split(" ")
            current_title = " ".join(parts[:6]).strip()
            current_company = " ".join(parts[6:12]).strip()

    return {
        "Status": status,
        "Applied For": applied_for,
        "Applied Age": applied_age,
        "Current Title (heuristic)": current_title,
        "Current Company (heuristic)": current_company,
    }


def autosize(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def main():
    if len(sys.argv) < 3:
        print("Usage: export_jobstreet_candidates_batch.py <input_json_path> <output_xlsx_path>")
        sys.exit(2)

    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2])

    data = json.loads(in_path.read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "candidates"

    headers = [
        "No",
        "Name",
        "Phone",
        "Status",
        "Applied For",
        "Applied Age",
        "Current Title (heuristic)",
        "Current Company (heuristic)",
        "Profile URL",
        "Source",
        "Raw Row Text",
        "Exported At",
    ]
    ws.append(headers)

    exported_at = datetime.now().isoformat(timespec="seconds")

    for i, r in enumerate(data, start=1):
        parsed = parse_row(r.get("rowText", ""))
        row = {
            "No": i,
            "Name": r.get("name", ""),
            "Phone": r.get("phone", ""),
            "Profile URL": r.get("profileUrl", ""),
            "Source": "JobStreet/SEEK 投递池",
            "Raw Row Text": r.get("rowText", ""),
            "Exported At": exported_at,
        }
        row.update(parsed)
        ws.append([row.get(h, "") for h in headers])

    autosize(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(str(out_path))


if __name__ == "__main__":
    main()
